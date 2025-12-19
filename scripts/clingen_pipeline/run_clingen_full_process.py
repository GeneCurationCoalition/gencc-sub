#!/usr/bin/env python3
"""
ClinGen Full Data Processing Pipeline

This master script runs the entire ClinGen data processing workflow:
1. Download and process ClinGen gene validity data
2. Download and process GCI Express data
3. Merge both datasets into a single target file
4. Export database submissions (requires PHP)
5. Compare target vs database
6. Generate merged submission files for upload

Run this script to execute the complete pipeline.
"""

import subprocess
import sys
import argparse
import shutil
import csv
from pathlib import Path
from datetime import datetime
import openpyxl
from openpyxl import load_workbook

# Configuration
SCRIPTS_DIR = Path(__file__).parent
DATA_DIR = SCRIPTS_DIR.parent.parent / "data" / "clingen"
TEMPLATE_DIR = SCRIPTS_DIR.parent.parent / "public" / "documents"
COMPARISON_DIR = DATA_DIR / "comparison"

# Extracted folders
GENE_VALIDITY_EXTRACTED = DATA_DIR / "gene_validity_extracted"
GCI_EXPRESS_EXTRACTED = DATA_DIR / "gci_express_extracted"

# Script paths
PROCESS_GENE_VALIDITY = SCRIPTS_DIR / "process_clingen_data.py"
PROCESS_GCI_EXPRESS = SCRIPTS_DIR / "process_gci_express_data.py"
CONVERT_GCI_EXPRESS = SCRIPTS_DIR / "convert_gci_express_to_tsv.py"
EXPORT_DATABASE = SCRIPTS_DIR / "export_clingen_submissions.php"
COMPARE_SUBMISSIONS = SCRIPTS_DIR / "compare_clingen_submissions.py"
GENERATE_MERGED = SCRIPTS_DIR / "generate_merged_submissions.py"

# Template and output files
GENCC_TEMPLATE = TEMPLATE_DIR / "GenCC Submission Spreadsheet.xlsx"
GENE_VALIDITY_FILE = DATA_DIR / "gene_validity_processed.tsv"
GCI_EXPRESS_FILE = DATA_DIR / "gci_express_processed.tsv"
COMBINED_TARGET_FILE = DATA_DIR / "gene_validity_with_gci_express.tsv"
DATABASE_EXPORT_FILE = DATA_DIR / "database_submissions_export.tsv"
ALL_CURRENT_CSV = COMPARISON_DIR / "all_current_submissions.csv"
DELETED_CSV = COMPARISON_DIR / "deleted_submissions_sgc.csv"
CHANGED_CSV = COMPARISON_DIR / "changed_submissions.csv"
ALL_CURRENT_EXCEL = COMPARISON_DIR / "all_current_submissions.xlsx"
DELETED_EXCEL = COMPARISON_DIR / "deleted_submissions.xlsx"
CHANGED_EXCEL = COMPARISON_DIR / "changed_submissions.xlsx"


class Colors:
    """Terminal colors for output"""
    GREEN = '\033[0;32m'
    YELLOW = '\033[1;33m'
    RED = '\033[0;31m'
    BLUE = '\033[0;34m'
    CYAN = '\033[0;36m'
    BOLD = '\033[1m'
    NC = '\033[0m'


def log_section(message: str) -> None:
    """Log a section header"""
    print(f"\n{Colors.BLUE}{Colors.BOLD}{'=' * 80}")
    print(f"{message}")
    print(f"{'=' * 80}{Colors.NC}\n")


def log_step(step_num: int, total_steps: int, message: str) -> None:
    """Log a step in the process"""
    print(f"{Colors.CYAN}[STEP {step_num}/{total_steps}]{Colors.NC} {message}")


def log_info(message: str) -> None:
    """Log an info message"""
    print(f"{Colors.GREEN}[INFO]{Colors.NC} {message}")


def log_warn(message: str) -> None:
    """Log a warning message"""
    print(f"{Colors.YELLOW}[WARN]{Colors.NC} {message}")


def log_error(message: str) -> None:
    """Log an error message"""
    print(f"{Colors.RED}[ERROR]{Colors.NC} {message}")


def run_command(cmd: list, description: str) -> bool:
    """Run a command and handle errors"""
    log_info(f"Running: {description}")
    log_info(f"Command: {' '.join(str(c) for c in cmd)}")

    try:
        result = subprocess.run(
            cmd,
            check=True,
            capture_output=False,
            text=True
        )
        log_info(f"✓ {description} completed successfully")
        return True
    except subprocess.CalledProcessError as e:
        log_error(f"✗ {description} failed with exit code {e.returncode}")
        return False
    except Exception as e:
        log_error(f"✗ {description} failed: {e}")
        return False


def merge_data_files() -> bool:
    """Merge gene validity and GCI Express data"""
    log_info("Merging gene validity and GCI Express data...")

    try:
        # Read gene validity data (includes header)
        with open(GENE_VALIDITY_FILE, 'r', encoding='utf-8') as f:
            gene_validity_lines = f.readlines()

        # Read GCI Express data (skip header)
        with open(GCI_EXPRESS_FILE, 'r', encoding='utf-8') as f:
            gci_express_lines = f.readlines()[1:]  # Skip header

        # Write combined file
        with open(COMBINED_TARGET_FILE, 'w', encoding='utf-8') as f:
            # Write all gene validity lines (includes header)
            f.writelines(gene_validity_lines)
            # Append GCI Express lines (without header)
            f.writelines(gci_express_lines)

        gene_validity_count = len(gene_validity_lines) - 1  # Exclude header
        gci_express_count = len(gci_express_lines)
        total_count = gene_validity_count + gci_express_count

        log_info(f"✓ Merged files successfully:")
        log_info(f"  - Gene Validity: {gene_validity_count} records")
        log_info(f"  - GCI Express: {gci_express_count} records")
        log_info(f"  - Combined Total: {total_count} records")
        log_info(f"  - Output: {COMBINED_TARGET_FILE}")

        return True

    except Exception as e:
        log_error(f"✗ Failed to merge data files: {e}")
        return False


def csv_to_excel(csv_file: Path, output_file: Path, sheet_name: str) -> bool:
    """
    Convert CSV file to Excel format using GenCC template

    Args:
        csv_file: Path to input CSV file
        output_file: Path to output Excel file
        sheet_name: Name for the data sheet

    Returns:
        True if successful, False otherwise
    """
    try:
        # Check if template exists
        if not GENCC_TEMPLATE.exists():
            log_error(f"Template not found: {GENCC_TEMPLATE}")
            return False

        # Check if CSV exists
        if not csv_file.exists():
            log_error(f"CSV file not found: {csv_file}")
            return False

        log_info(f"Converting {csv_file.name} to Excel format...")

        # Load the template
        wb = load_workbook(GENCC_TEMPLATE)
        ws = wb['Data']

        # Rename the sheet
        ws.title = sheet_name

        # Read CSV data (skip header, BOM handled automatically)
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader)  # Skip CSV header
            csv_data = list(reader)

        # CSV column mapping to Excel columns
        # CSV: SGC ID, Submission ID, HGNC ID, Gene Symbol, Disease ID (MONDO), Disease Name,
        #      Mode of Inheritance ID, Mode of Inheritance Name, Submitter ID, Submitter Name,
        #      Classification ID, Classification Name, Report Date, Public Report URL, Notes,
        #      PubMed IDs, Assertion Criteria URL
        # Excel columns (A=1): sgc_id(1), local_key(2), hgnc_id(3), hgnc_symbol(4), disease_id(5),
        #      disease_name(6), moi_id(7), moi_name(8), submitter_id(9), submitter_name(10),
        #      classification_id(11), classification_name(12), report_date(13),
        #      public_report_url(14), notes(15), pmids(16), assertion_criteria_url(17)

        # Helper function to clean illegal characters for Excel
        def clean_for_excel(value):
            """Remove illegal characters that Excel doesn't accept"""
            if not isinstance(value, str):
                return value
            # Remove control characters except tab, newline, and carriage return
            import re
            return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', value)

        # Write data starting at row 13
        start_row = 13
        for row_idx, csv_row in enumerate(csv_data, start=start_row):
            # Map CSV columns to Excel columns, cleaning values
            ws.cell(row=row_idx, column=1, value=clean_for_excel(csv_row[0]))   # SGC ID
            ws.cell(row=row_idx, column=2, value=clean_for_excel(csv_row[1]))   # Submission ID (local_key)
            ws.cell(row=row_idx, column=3, value=clean_for_excel(csv_row[2]))   # HGNC ID
            ws.cell(row=row_idx, column=4, value=clean_for_excel(csv_row[3]))   # Gene Symbol
            ws.cell(row=row_idx, column=5, value=clean_for_excel(csv_row[4]))   # Disease ID
            ws.cell(row=row_idx, column=6, value=clean_for_excel(csv_row[5]))   # Disease Name
            ws.cell(row=row_idx, column=7, value=clean_for_excel(csv_row[6]))   # MOI ID
            ws.cell(row=row_idx, column=8, value=clean_for_excel(csv_row[7]))   # MOI Name
            ws.cell(row=row_idx, column=9, value=clean_for_excel(csv_row[8]))   # Submitter ID
            ws.cell(row=row_idx, column=10, value=clean_for_excel(csv_row[9]))  # Submitter Name
            ws.cell(row=row_idx, column=11, value=clean_for_excel(csv_row[10])) # Classification ID
            ws.cell(row=row_idx, column=12, value=clean_for_excel(csv_row[11])) # Classification Name
            ws.cell(row=row_idx, column=13, value=clean_for_excel(csv_row[12])) # Report Date
            ws.cell(row=row_idx, column=14, value=clean_for_excel(csv_row[13])) # Public Report URL
            ws.cell(row=row_idx, column=15, value=clean_for_excel(csv_row[14])) # Notes
            ws.cell(row=row_idx, column=16, value=clean_for_excel(csv_row[15])) # PubMed IDs
            ws.cell(row=row_idx, column=17, value=clean_for_excel(csv_row[16])) # Assertion Criteria URL

            # Handle assertion_criteria_url if it exists (column 18 in CSV, column 18 in Excel)
            if len(csv_row) > 17:
                ws.cell(row=row_idx, column=18, value=clean_for_excel(csv_row[17])) # Assertion Criteria URL (actual)

        # Delete any extra rows beyond our data to prevent upload issues
        # Calculate the last row with actual data
        last_data_row = start_row + len(csv_data) - 1
        rows_to_delete = ws.max_row - last_data_row
        if rows_to_delete > 0:
            ws.delete_rows(last_data_row + 1, rows_to_delete)
            log_info(f"  - Deleted {rows_to_delete} empty rows from template")
            log_info(f"  - Final row count: {ws.max_row}")

        # Save the workbook
        wb.save(output_file)

        log_info(f"✓ Created Excel file: {output_file}")
        log_info(f"  - Sheet name: {sheet_name}")
        log_info(f"  - Records: {len(csv_data)}")

        return True

    except Exception as e:
        log_error(f"✗ Failed to convert CSV to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def cleanup_extracted_folders(keep_extracted: bool = False) -> None:
    """
    Clean up extracted JSON folders

    Args:
        keep_extracted: If True, keep the extracted folders; if False, delete them
    """
    if keep_extracted:
        log_info("Keeping extracted folders as requested")
        return

    folders_to_remove = [GENE_VALIDITY_EXTRACTED, GCI_EXPRESS_EXTRACTED]

    for folder in folders_to_remove:
        if folder.exists():
            try:
                shutil.rmtree(folder)
                log_info(f"✓ Removed extracted folder: {folder.name}")
            except Exception as e:
                log_warn(f"Failed to remove {folder.name}: {e}")
        else:
            log_info(f"  No extracted folder to remove: {folder.name}")


def main():
    """Main execution"""
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='ClinGen Full Data Processing Pipeline',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python3 run_clingen_full_process.py              # Run pipeline with cleanup
  python3 run_clingen_full_process.py --keep-extracted  # Keep extracted folders
        """
    )
    parser.add_argument(
        '--keep-extracted',
        action='store_true',
        help='Keep extracted JSON folders after processing (default: delete them)'
    )

    args = parser.parse_args()

    start_time = datetime.now()

    log_section("ClinGen Full Data Processing Pipeline")
    log_info(f"Started at: {start_time.strftime('%Y-%m-%d %H:%M:%S')}")

    if args.keep_extracted:
        log_info("Mode: Keeping extracted folders after completion")

    total_steps = 8
    current_step = 0

    # Step 1: Process ClinGen gene validity data
    current_step += 1
    log_step(current_step, total_steps, "Processing ClinGen gene validity data")
    if not run_command(
        ["python3", str(PROCESS_GENE_VALIDITY)],
        "ClinGen gene validity processing"
    ):
        log_error("Pipeline aborted at Step 1")
        sys.exit(1)

    # Step 2: Download GCI Express data
    current_step += 1
    log_step(current_step, total_steps, "Downloading and extracting GCI Express data")
    if not run_command(
        ["python3", str(PROCESS_GCI_EXPRESS)],
        "GCI Express download and extraction"
    ):
        log_error("Pipeline aborted at Step 2")
        sys.exit(1)

    # Step 3: Convert GCI Express to TSV
    current_step += 1
    log_step(current_step, total_steps, "Converting GCI Express data to TSV format")
    if not run_command(
        ["python3", str(CONVERT_GCI_EXPRESS)],
        "GCI Express TSV conversion"
    ):
        log_error("Pipeline aborted at Step 3")
        sys.exit(1)

    # Step 4: Merge gene validity and GCI Express data
    current_step += 1
    log_step(current_step, total_steps, "Merging gene validity and GCI Express data")
    if not merge_data_files():
        log_error("Pipeline aborted at Step 4")
        sys.exit(1)

    # Step 5: Export database submissions
    current_step += 1
    log_step(current_step, total_steps, "Exporting database submissions")
    if not run_command(
        ["php", str(EXPORT_DATABASE)],
        "Database submissions export"
    ):
        log_error("Pipeline aborted at Step 5")
        sys.exit(1)

    # Step 6: Compare target vs database
    current_step += 1
    log_step(current_step, total_steps, "Comparing target data with database")
    if not run_command(
        ["python3", str(COMPARE_SUBMISSIONS)],
        "Target vs database comparison"
    ):
        log_error("Pipeline aborted at Step 6")
        sys.exit(1)

    # Step 7: Generate merged submission files
    current_step += 1
    log_step(current_step, total_steps, "Generating merged submission files for upload")
    if not run_command(
        ["python3", str(GENERATE_MERGED)],
        "Merged submission file generation"
    ):
        log_error("Pipeline aborted at Step 7")
        sys.exit(1)

    # Step 8: Convert CSV files to Excel format
    current_step += 1
    log_step(current_step, total_steps, "Converting CSV files to Excel format")

    log_info("Creating Excel file for all current submissions...")
    if not csv_to_excel(ALL_CURRENT_CSV, ALL_CURRENT_EXCEL, "All Current Submissions"):
        log_error("Failed to create all current submissions Excel file")
        sys.exit(1)

    log_info("Creating Excel file for deleted submissions...")
    if not csv_to_excel(DELETED_CSV, DELETED_EXCEL, "Deleted Submissions"):
        log_error("Failed to create deleted submissions Excel file")
        sys.exit(1)

    log_info("Creating Excel file for changed submissions...")
    if not csv_to_excel(CHANGED_CSV, CHANGED_EXCEL, "Changed Submissions"):
        log_error("Failed to create changed submissions Excel file")
        sys.exit(1)

    # Cleanup extracted folders
    log_info("")
    log_info("Cleaning up extracted folders...")
    cleanup_extracted_folders(keep_extracted=args.keep_extracted)

    # Complete
    end_time = datetime.now()
    duration = end_time - start_time

    log_section("Pipeline Completed Successfully!")

    print(f"{Colors.GREEN}{Colors.BOLD}✓ All steps completed successfully!{Colors.NC}\n")
    print(f"Started:  {start_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Finished: {end_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"Duration: {duration}\n")

    print(f"{Colors.CYAN}Output Files:{Colors.NC}")
    print(f"  1. Combined Target:               {COMBINED_TARGET_FILE}")
    print(f"  2. Database Export:               {DATABASE_EXPORT_FILE}")
    print(f"  3. Comparison Reports:            {DATA_DIR}/comparison/")
    print(f"  4. All Current Submissions CSV:   {ALL_CURRENT_CSV}")
    print(f"  5. Changed Submissions CSV:       {CHANGED_CSV}")
    print(f"  6. Deleted Submissions CSV:       {DELETED_CSV}")
    print(f"  7. All Current Submissions XLS:   {ALL_CURRENT_EXCEL}")
    print(f"  8. Changed Submissions XLS:       {CHANGED_EXCEL}")
    print(f"  9. Deleted Submissions XLS:       {DELETED_EXCEL}")
    print()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log_warn("\nPipeline interrupted by user")
        sys.exit(1)
    except Exception as e:
        log_error(f"Pipeline failed with unexpected error: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
