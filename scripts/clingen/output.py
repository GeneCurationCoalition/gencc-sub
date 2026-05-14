"""
Output generation module for ClinGen Pipeline

Generates the final CSV and Excel output files for upload.
"""

import csv
import re
from pathlib import Path
from typing import Dict, List

from .config import (
    OutputConfig, DOWNLOAD_COLUMNS, TARGET_TSV_FIELDS, DATABASE_TSV_FIELDS,
    COMPARISON_DIR, TEMPLATE_DIR
)
from .logging import log_info, log_warn, log_error, log_success, log_section
from .models import Submission, ComparisonResult
from .comparison import (
    load_target_submissions, load_database_submissions,
    has_changes
)

try:
    from openpyxl import load_workbook
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def generate_csv_outputs(
    result: ComparisonResult,
    target: Dict[str, Submission],
    database: Dict[str, Submission],
    output_config: OutputConfig = None
) -> None:
    """
    Generate CSV output files.

    Creates:
    - all_current_submissions.csv: All target records
    - changed_submissions.csv: Records with changes + new
    - deleted_submissions_sgc.csv: Records to unpublish

    Args:
        result: ComparisonResult from comparison.
        target: Dictionary of target submissions.
        database: Dictionary of database submissions.
        output_config: Output configuration.
    """
    if output_config is None:
        output_config = OutputConfig()

    log_section("Generating CSV Output Files")

    # Ensure directory exists
    COMPARISON_DIR.mkdir(parents=True, exist_ok=True)

    # Generate all current submissions
    all_records = []
    unchanged_count = 0

    for local_id, sub in target.items():
        # Determine action and SGC ID
        if local_id in result.new_submissions:
            action = 'N'
            sgc_id = ''
        elif local_id in result.matched_by_id:
            action = 'R'
            sgc_id = result.sgc_mapping.get(local_id, '')
        elif local_id in result.gdm_id_mapping:
            action = 'R'
            sgc_id = result.sgc_mapping.get(local_id, '')
        else:
            action = 'R'
            sgc_id = result.sgc_mapping.get(local_id, '')
            unchanged_count += 1

        # Create submission with SGC ID
        sub_with_sgc = Submission(
            sgc_id=sgc_id,
            local_id=sub.local_id,
            gene_id=sub.gene_id,
            disease_id=sub.disease_id,
            mode_of_inheritance=sub.mode_of_inheritance,
            submitter_id=sub.submitter_id,
            classification_id=sub.classification_id,
            classification=sub.classification,
            report_date=sub.report_date,
            public_report_url=sub.public_report_url,
            notes=sub.notes,
            pubmed_ids=sub.pubmed_ids,
            assertion_criteria_url=sub.assertion_criteria_url
        )
        all_records.append(sub_with_sgc.to_download_row(action))

    # Write all current submissions
    with open(output_config.all_current_csv, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(DOWNLOAD_COLUMNS)
        writer.writerows(all_records)

    log_success(f"All current submissions: {output_config.all_current_csv}")
    log_info(f"  Total records: {len(all_records)}")

    # Generate changed submissions (changes + republishes + new)
    changed_records = []

    # Add records matched by ID with changes or being republished.
    # Exclude GDM-matched records — they are handled in the GDM loop below.
    gdm_matched_ids = {m.target_submission.local_id for m in result.matched_by_gdm}
    changed_or_republished = (
        (result.matched_by_id_changed | result.republished) - gdm_matched_ids
    )
    for local_id in changed_or_republished:
        if local_id in target:
            sub = target[local_id]
            sgc_id = result.sgc_mapping.get(local_id, '')

            sub_with_sgc = Submission(
                sgc_id=sgc_id,
                local_id=sub.local_id,
                gene_id=sub.gene_id,
                disease_id=sub.disease_id,
                mode_of_inheritance=sub.mode_of_inheritance,
                submitter_id=sub.submitter_id,
                classification_id=sub.classification_id,
                classification=sub.classification,
                report_date=sub.report_date,
                public_report_url=sub.public_report_url,
                notes=sub.notes,
                pubmed_ids=sub.pubmed_ids,
                assertion_criteria_url=sub.assertion_criteria_url
            )
            changed_records.append(sub_with_sgc.to_download_row('R'))

    # Add records matched by GDM (these are always considered changes)
    for match in result.matched_by_gdm:
        sub = match.target_submission
        sgc_id = match.database_submission.sgc_id

        sub_with_sgc = Submission(
            sgc_id=sgc_id,
            local_id=sub.local_id,
            gene_id=sub.gene_id,
            disease_id=sub.disease_id,
            mode_of_inheritance=sub.mode_of_inheritance,
            submitter_id=sub.submitter_id,
            classification_id=sub.classification_id,
            classification=sub.classification,
            report_date=sub.report_date,
            public_report_url=sub.public_report_url,
            notes=sub.notes,
            pubmed_ids=sub.pubmed_ids,
            assertion_criteria_url=sub.assertion_criteria_url
        )
        changed_records.append(sub_with_sgc.to_download_row('R'))

    # Add new submissions
    for local_id, sub in result.new_submissions.items():
        changed_records.append(sub.to_download_row('N'))

    # Add deleted submissions (action = U) into the changed file
    deleted_count = 0
    for local_id, sub in result.deleted_submissions.items():
        changed_records.append(sub.to_download_row('U'))
        deleted_count += 1

    # Write changed submissions (includes modifications, republishes, new, and deletes)
    with open(output_config.changed_csv, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.writer(f)
        writer.writerow(DOWNLOAD_COLUMNS)
        writer.writerows(changed_records)

    log_success(f"Changed submissions: {output_config.changed_csv}")
    log_info(f"  Total records: {len(changed_records)}")
    log_info(f"    - Modified: {len(result.matched_by_id_changed)}")
    log_info(f"    - Republished: {len(result.republished)}")
    log_info(f"    - GDM matched: {len(result.matched_by_gdm)}")
    log_info(f"    - New: {len(result.new_submissions)}")
    log_info(f"    - Deleted: {deleted_count}")


def csv_to_excel(
    csv_file: Path,
    output_file: Path,
    sheet_name: str,
    template_path: Path = None
) -> bool:
    """
    Convert CSV file to Excel format using GenCC template.

    Args:
        csv_file: Path to input CSV file.
        output_file: Path to output Excel file.
        sheet_name: Name for the data sheet.
        template_path: Path to Excel template. Uses default if None.

    Returns:
        True if successful, False otherwise.
    """
    if not HAS_OPENPYXL:
        log_error("openpyxl is required for Excel export. Install with: pip install openpyxl")
        return False

    if template_path is None:
        template_path = TEMPLATE_DIR / "GenCC Submission Spreadsheet.xlsx"

    if not template_path.exists():
        log_error(f"Template not found: {template_path}")
        return False

    if not csv_file.exists():
        log_error(f"CSV file not found: {csv_file}")
        return False

    log_info(f"Converting {csv_file.name} to Excel format...")

    try:
        # Load the template
        wb = load_workbook(template_path)
        ws = wb['Data']
        ws.title = sheet_name

        # Read CSV data (skip header, BOM handled by utf-8-sig)
        with open(csv_file, 'r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            csv_data = list(reader)

        # Helper to clean illegal characters
        def clean_for_excel(value):
            if not isinstance(value, str):
                return value
            return re.sub(r'[\x00-\x08\x0B-\x0C\x0E-\x1F\x7F-\x9F]', '', value)

        # Write data starting at row 13
        start_row = 13
        for row_idx, csv_row in enumerate(csv_data, start=start_row):
            for col_idx, value in enumerate(csv_row, start=1):
                ws.cell(row=row_idx, column=col_idx, value=clean_for_excel(value))

        # Delete empty rows beyond data
        last_data_row = start_row + len(csv_data) - 1
        rows_to_delete = ws.max_row - last_data_row
        if rows_to_delete > 0:
            ws.delete_rows(last_data_row + 1, rows_to_delete)

        wb.save(output_file)

        log_success(f"Created Excel file: {output_file}")
        log_info(f"  Sheet name: {sheet_name}")
        log_info(f"  Records: {len(csv_data)}")

        return True

    except Exception as e:
        log_error(f"Failed to convert CSV to Excel: {e}")
        import traceback
        traceback.print_exc()
        return False


def generate_excel_outputs(output_config: OutputConfig = None) -> bool:
    """
    Generate Excel output files from CSVs.

    Args:
        output_config: Output configuration.

    Returns:
        True if all files created successfully.
    """
    if output_config is None:
        output_config = OutputConfig()

    log_section("Generating Excel Output Files")

    # Hard fail upfront if openpyxl is missing — don't silently continue and
    # leave behind stale XLSX files from a previous run.
    if not HAS_OPENPYXL:
        log_error(
            "openpyxl is not installed. Excel output cannot be generated. "
            "Install dependencies with: pip install -r scripts/requirements.txt"
        )
        return False

    success = True

    # All current submissions
    if not csv_to_excel(
        output_config.all_current_csv,
        output_config.all_current_xlsx,
        "All Current Submissions"
    ):
        success = False

    # Changed submissions (includes modifications, republishes, new, and deletes)
    if not csv_to_excel(
        output_config.changed_csv,
        output_config.changed_xlsx,
        "Changed Submissions"
    ):
        success = False

    return success


def run_output_generation(output_config: OutputConfig = None) -> bool:
    """
    Run the complete output generation workflow.

    Loads comparison results and generates CSV/Excel outputs.

    Args:
        output_config: Output configuration.

    Returns:
        True if successful.
    """
    if output_config is None:
        output_config = OutputConfig()

    # Load data
    log_info("Loading data for output generation...")

    target = load_target_submissions(output_config.combined_target)
    database = load_database_submissions(output_config.database_export)

    # Re-run comparison to get structured result
    from .comparison import compare_submissions, load_unpublished_local_ids

    unpublished_ids = load_unpublished_local_ids(output_config.unpublished_ids)
    result = compare_submissions(target, database, unpublished_ids)

    # Generate outputs
    generate_csv_outputs(result, target, database, output_config)
    generate_excel_outputs(output_config)

    return True
